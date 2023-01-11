import sys                                              # txt 파일 저장하려 열기
import datetime                                         # 오늘 시간 구하기 / 일주일 단위 구하기
import csv
import requests                                         # 뉴스 주소 검색어로 만들기
from bs4 import BeautifulSoup                           # crawling 만들기
import pyautogui                                        # 검색어 입력
import openpyxl                                         # 엑셀 파일 만들기
from openpyxl.styles.fonts import Font                 # 엑셀 셀서식 지정
# from openpyxl import Workbook
# import xlsxwriter                                       # 엑셀에 하이퍼링크 적용
# import pandas as pd                                     # 엑셀 데이터 추출

today_org = datetime.date.today()                                                               # 검색 기사 시작일 ~ 종료일 지정
week_ago_org = today_org - datetime.timedelta(days=6)
today = today_org.strftime("%Y.%m.%d")
week_ago = week_ago_org.strftime("%Y.%m.%d")
# print(week_ago.strftime("%Y.%m.%d"))

keyword = pyautogui.prompt("검색어를 입력하세요.")                                                # 기사 검색어 입력 팝업

sys.stdout = open(f'이번주기사_크롤링_{keyword}_{today}.txt', 'w', encoding='UTF8')               # txt 파일 만들기
# sys.stdout = open('print_저장하기.csv', 'w', encoding='UTF8')

print(f"기사 시작일\t{week_ago}")
print(f"기사 종료일\t{today}\n")
print("\t기사 제목\t링크 주소")

for i in range(1, 100, 10):
#    response = requests.get(f"https://search.naver.com/search.naver?where=news&sm=tab_pge&query={keyword}&sort=0&photo=0&field=0&pd=1&ds=2022.08.31&de=2022.09.07&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so:r,p:1w,a:all&start={i}")
    response = requests.get(f"https://search.naver.com/search.naver?where=news&sm=tab_pge&query={keyword}&sort=0&photo=0&field=0&pd=1&ds={week_ago}&de={today}&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so:r,p:1w,a:all&start={i}")
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    links = soup.select(".news_tit")                # 결과는 리스트

    # for item in data:
    #         product_name = item.select_one('div.card-body h4.card-text') # 상품명 크롤링
    #         product_date = item.select_one('div.wrapfooter span.post-date') # 상품 등록 날짜 크롤링
    #         product_info = [product_name.get_text().strip(), product_date.get_text().strip()] # 리스트에 담기
    #         product_lists.append(product_info) # 전역변수로 선언한 리스트에 추가(이중 리스트 : 리스트안에 상품별로 리스트 생성)

    for link in links:
        title = link.text                   # 태그 안에 텍스트요소를 가져온다   # 기사제목 크롤링
        url = link.attrs['href']            # href의 속성값을 가져온다         # 기사주소 크롤링
        # print(f"\t{title}\t{url}")          # 기사제목 (탭) 기사주소
        print(f"\t{title}\t=HYPERLINK(\"{url}\",\"{url}\")")
                
sys.stdout.close()

input_file = f'이번주기사_크롤링_{keyword}_{today}.txt'                         # txt 파일을 xlsx 파일로 변환
output_file = f'이번주기사_크롤링_{keyword}_{today}.xlsx'

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

with open(input_file, 'r', encoding="UTF8") as data:
    reader = csv.reader(data, delimiter='\t')           # 구분자 = 탭(\t)
    for row in reader:
        ws.append(row)

    # 셀 서식
    ws.column_dimensions['A'].width = 11    #A열의 너비 11
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 120
    ws['A1'].font = Font(bold = True,       # 기사 시작일
                        italic = True)
    ws['B1'].font = Font(bold = True,       # 기사 시작일 날짜
                        italic = True)
    ws['A2'].font = Font(bold = True,       # 기사 종료일
                        italic = True)
    ws['B2'].font = Font(bold = True,       # 기사 종료일 날짜
                        italic = True)
    ws['B4'].font = Font(bold = True,       # 기사 제목
                        italic = True)
    ws['C4'].font = Font(bold = True,       # 링크 주소
                        italic = True)
    for j in range(1, 101, 1):
        ws.cell(column = 3, row = j + 4).font = Font(italic = True,
                                                    size = 9,
#                                                    underline = True
                                                    color = "0000FF")

        # ws[f'C{j}+4'].font = Font(italic = True,
        #                          color = "0000FF",
        #                          underline = True)

    # 하이퍼링크 양식 생성
    # workbook = xlsxwriter.Workbook(f'이번주기사_크롤링_{keyword}_{today}.xlsx')         # 엑셀 파일 생성
    # worksheet = wb.active                                                              # 활성화된 엑셀 시트 선택
    # worksheet = workbook.add_worksheet('Sheet1')                                      # 엑셀 시트 생성

    # link_format = workbook.add_format({
    #     'font_color': 'blue',
    #     'underline': 1,
    #     'italic': 1,
    #     'font_size': 9,
    # })

    # worksheet.write_url('C5', '')
#    worksheet.write_url('C5', '', link_format)

    ##################################################

    # for j in range(1, 10, 1):
        # cell_content(f'{j}'+4) = w1.cell(row=f'{j}'+4, column=3).value
        # j = j+1

        # worksheet.write_url(f'C{j+4}', link_format)
    # worksheet.write_url('A1', 'http://www.python.org/')       # 하이퍼링크 적용

    # ws.cell(row=5, column=3).value = '=HYPERLINK("{}", "{}")'.format(link, "Link Name")

wb.save(output_file)
